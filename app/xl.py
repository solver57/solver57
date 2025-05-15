import openpyxl
from openpyxl.styles import PatternFill
import colorsys


def solution2matrix(s):
    t = {}
    for r in s:
        ind = r[-1][:1]
        color = ind if r[-1] != 'exclude' else '0'
        for w in r[:-1]:
            t[tuple(map(int, w.split('_')))] = color
    h, w = map(lambda x: max(x) + 1, zip(*list(t)))
    m = [[t[(i, j)] for j in range(w)] for i in range(h)]
    return m


def create_colored_excel(matrices, filename='output.xlsx'):
    wb = openpyxl.Workbook()
    ws = wb.active

    # Получаем все уникальные буквы (исключая '0')
    unique_letters = sorted({cell for matrix in matrices for row in matrix for cell in row if cell != '0'})

    # Создаем палитру цветов
    palette = []
    num_colors = len(unique_letters)
    for i in range(num_colors):
        hue = i / num_colors  # Равномерное распределение оттенков
        saturation = [0.5, 0.9, 0.9, 0.5][i % 4]  # Разная насыщенность
        lightness = [0.4, 0.9][i % 2]  # Разная яркость
        rgb = colorsys.hls_to_rgb(hue, lightness, saturation)
        hex_color = f"{int(rgb[0] * 255):02X}{int(rgb[1] * 255):02X}{int(rgb[2] * 255):02X}"
        palette.append(hex_color)

    # Сопоставляем буквы и цвета
    color_map = {letter: PatternFill(start_color=color, end_color=color, fill_type="solid")
                 for letter, color in zip(unique_letters, palette)}

    # Размещаем матрицы на листе по N в ряд
    N_IN_ROW = 6
    row_offset = 1
    for i, matrix in enumerate(matrices):
        # Определяем позицию матрицы
        col_offset = (i % N_IN_ROW) * (len(matrix[0]) + 2) + 1
        if i % N_IN_ROW == 0 and i != 0:
            row_offset += len(matrix) + 2

        # Закрашиваем ячейки
        for r, row in enumerate(matrix, start=row_offset):
            for c, cell in enumerate(row, start=col_offset):
                if cell != '0':
                    ws.cell(row=r, column=c).fill = color_map[cell]

    # Делаем ячейки квадратными
    for col in ws.columns:
        ws.column_dimensions[
            col[0].column_letter].width = 3  # Ширина столбца = 3 (примерно соответствует высоте строки)

    # Сохраняем файл
    wb.save(filename)

def rotate_matrix(matrix):
    """Поворачивает матрицу на 90 градусов по часовой стрелке."""
    return [list(row) for row in zip(*matrix[::-1])]

def reflect_matrix(matrix):
    """Отражает матрицу по горизонтали."""
    return [row[::-1] for row in matrix]

def generate_all_variants(matrix):
    """Генерирует все 8 вариантов матрицы (повороты и отражения)."""
    variants = []
    current = matrix
    for _ in range(4):  # 4 поворота
        variants.append(current)
        variants.append(reflect_matrix(current))  # Отражение
        current = rotate_matrix(current)  # Поворот на 90 градусов
    return variants

def find_unique_matrices(matrices):
    """Возвращает список уникальных матриц с точностью до поворотов и отражений."""
    unique_matrices = []
    for matrix in matrices:
        # Генерируем все варианты текущей матрицы
        variants = generate_all_variants(matrix)
        # Проверяем, есть ли хотя бы один вариант в списке уникальных
        is_unique = True
        for variant in variants:
            if variant in unique_matrices:
                is_unique = False
                break
        # Если матрица уникальна, добавляем её в список
        if is_unique:
            unique_matrices.append(matrix)
    return unique_matrices



# Пример использования
if __name__ == "__main__":
    # Пример использования
    matrices = [
        [['X', 'X', 'X', 'S', 'P', 'P', 'P', 'P'],
         ['R', 'X', 'S', 'S', 'S', 'Q', 'Q', 'P'],
         ['R', 'X', 'U', 'S', 'Q', 'Q', 'Z', 'Z'],
         ['R', 'U', 'U', '0', '0', 'Q', 'Z', 'Z'],
         ['R', 'Y', 'U', '0', '0', 'V', 'V', 'Z'],
         ['R', 'Y', 'U', 'W', 'O', 'O', 'V', 'T'],
         ['Y', 'Y', 'W', 'W', 'O', 'V', 'V', 'T'],
         ['Y', 'W', 'W', 'O', 'O', 'T', 'T', 'T']],

        [['X', 'X', 'X', 'S', 'P', 'P', 'P', 'P'],
         ['R', 'X', 'S', 'S', 'S', 'Q', 'Q', 'P'],
         ['R', 'X', 'U', 'S', 'Q', 'Q', 'Z', 'Z'],
         ['R', 'U', 'U', '0', '0', 'Q', 'Z', 'Z'],
         ['R', 'Y', 'U', '0', '0', 'V', 'V', 'Z'],
         ['R', 'Y', 'U', 'W', 'O', 'O', 'V', 'T'],
         ['Y', 'Y', 'W', 'W', 'O', 'V', 'V', 'T'],
         ['Y', 'W', 'W', 'O', 'O', 'T', 'T', 'T']],

        [['X', 'X', 'X', 'S', 'P', 'P', 'P', 'P'],
         ['R', 'X', 'S', 'S', 'S', 'Q', 'Q', 'P'],
         ['R', 'X', 'U', 'S', 'Q', 'Q', 'Z', 'Z'],
         ['R', 'U', 'U', '0', '0', 'Q', 'Z', 'Z'],
         ['R', 'Y', 'U', '0', '0', 'V', 'V', 'Z'],
         ['R', 'Y', 'U', 'W', 'O', 'O', 'V', 'T'],
         ['Y', 'Y', 'W', 'W', 'O', 'V', 'V', 'T'],
         ['Y', 'W', 'W', 'O', 'O', 'T', 'T', 'T']],
    ]

    unique_matrices = find_unique_matrices(matrices)
    print(f"Уникальных матриц: {len(unique_matrices)}")

    matrices = [
        [['X', 'X', 'X', 'S', 'P', 'P', 'P', 'P'],
         ['R', 'X', 'S', 'S', 'S', 'Q', 'Q', 'P'],
         ['R', 'X', 'U', 'S', 'Q', 'Q', 'Z', 'Z'],
         ['R', 'U', 'U', '0', '0', 'Q', 'Z', 'Z'],
         ['R', 'Y', 'U', '0', '0', 'V', 'V', 'Z'],
         ['R', 'Y', 'U', 'W', 'O', 'O', 'V', 'T'],
         ['Y', 'Y', 'W', 'W', 'O', 'V', 'V', 'T'],
         ['Y', 'W', 'W', 'O', 'O', 'T', 'T', 'T']],

        [['X', 'X', 'X', 'S', 'P', 'P', 'P', 'P'],
         ['R', 'X', 'S', 'S', 'S', 'Q', 'Q', 'P'],
         ['R', 'X', 'U', 'S', 'Q', 'Q', 'Z', 'Z'],
         ['R', 'U', 'U', '0', '0', 'Q', 'Z', 'Z'],
         ['R', 'Y', 'U', '0', '0', 'V', 'V', 'Z'],
         ['R', 'Y', 'U', 'W', 'O', 'O', 'V', 'T'],
         ['Y', 'Y', 'W', 'W', 'O', 'V', 'V', 'T'],
         ['Y', 'W', 'W', 'O', 'O', 'T', 'T', 'T']],

        [['X', 'X', 'X', 'S', 'P', 'P', 'P', 'P'],
         ['R', 'X', 'S', 'S', 'S', 'Q', 'Q', 'P'],
         ['R', 'X', 'U', 'S', 'Q', 'Q', 'Z', 'Z'],
         ['R', 'U', 'U', '0', '0', 'Q', 'Z', 'Z'],
         ['R', 'Y', 'U', '0', '0', 'V', 'V', 'Z'],
         ['R', 'Y', 'U', 'W', 'O', 'O', 'V', 'T'],
         ['Y', 'Y', 'W', 'W', 'O', 'V', 'V', 'T'],
         ['Y', 'W', 'W', 'O', 'O', 'T', 'T', 'T']],

        [['X', 'X', 'X', 'S', 'P', 'P', 'P', 'P'],
         ['R', 'X', 'S', 'S', 'S', 'Q', 'Q', 'P'],
         ['R', 'X', 'U', 'S', 'Q', 'Q', 'Z', 'Z'],
         ['R', 'U', 'U', '0', '0', 'Q', 'Z', 'Z'],
         ['R', 'Y', 'U', '0', '0', 'V', 'V', 'Z'],
         ['R', 'Y', 'U', 'W', 'O', 'O', 'V', 'T'],
         ['Y', 'Y', 'W', 'W', 'O', 'V', 'V', 'T'],
         ['Y', 'W', 'W', 'O', 'O', 'T', 'T', 'T']],
        [['X', 'X', 'X', 'S', 'P', 'P', 'P', 'P'],
         ['R', 'X', 'S', 'S', 'S', 'Q', 'Q', 'P'],
         ['R', 'X', 'U', 'S', 'Q', 'Q', 'Z', 'Z'],
         ['R', 'U', 'U', '0', '0', 'Q', 'Z', 'Z'],
         ['R', 'Y', 'U', '0', '0', 'V', 'V', 'Z'],
         ['R', 'Y', 'U', 'W', 'O', 'O', 'V', 'T'],
         ['Y', 'Y', 'W', 'W', 'O', 'V', 'V', 'T'],
         ['Y', 'W', 'W', 'O', 'O', 'T', 'T', 'T']],
    ]

    create_colored_excel(matrices)
